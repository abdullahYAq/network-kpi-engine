import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
def create_empty_kpi_template(excel_path,header_name):
    kpi_sheet_name = "New KPI Insertion"
    kpi_excel_data = pd.DataFrame(data=[],columns=header_name)
    kpi_excel_data.to_excel(excel_path,sheet_name=kpi_sheet_name,index=False)
    workbook = openpyxl.load_workbook(excel_path)
    sheet_kpi_opxl = workbook[kpi_sheet_name]
    row_num = 100
    validation_kpi_type = DataValidation(type="list",formula1='"ratio (num/den),expression"')
    sheet_kpi_opxl.add_data_validation(validation_kpi_type)
    validation_kpi_type.add(f"B2:B{row_num}")
    validation_tech = DataValidation(type="list",formula1='"NSANR,LTE,UMTS,GSM"')
    sheet_kpi_opxl.add_data_validation(validation_tech)
    validation_tech.add(f"H2:H{row_num}")
    sheet_kpi_opxl.conditional_formatting.add(f"C2:C{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="expression"'], fill=openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type = "solid")))
    sheet_kpi_opxl.conditional_formatting.add(f"C2:C{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="ratio (num/den)"'], fill=openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type = "solid")))
    sheet_kpi_opxl.conditional_formatting.add(f"D2:D{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="expression"'], fill=openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type = "solid")))
    sheet_kpi_opxl.conditional_formatting.add(f"D2:D{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="ratio (num/den)"'], fill=openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type = "solid")))
    sheet_kpi_opxl.conditional_formatting.add(f"E2:E{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="expression"'], fill=openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type = "solid")))
    sheet_kpi_opxl.conditional_formatting.add(f"E2:E{row_num}", openpyxl.formatting.rule.FormulaRule(formula = ['=$B2="ratio (num/den)"'], fill=openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type = "solid")))
    sheet_kpi_opxl['C1'].comment = Comment("Write formula like: A + B - C\nUsed only when kpi_type = ratio",
    "System")
    sheet_kpi_opxl['D1'].comment = Comment("Write formula like: A + B + C\nUsed only when kpi_type = ratio",
    "System")
    sheet_kpi_opxl['E1'].comment = Comment("Write formula like: A + B - C\nUsed only when kpi_type = expression",
    "System")
    sheet_kpi_opxl['F1'].comment = Comment("if you want to multiply the final result by a number, write it here, \notherwise leave it empty and it will default to 1\n note: the multiplier will be applied to the final result of the formula, whether it's a ratio or an expression",
    "System")
    auto_adjust_column_width(sheet_kpi_opxl)
    sheet_kpi_opxl.freeze_panes = "A2"
    workbook.save(excel_path)
def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column  # رقم العمود
        
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # padding بسيط
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
def write_errors_report(excel_path, errors_df, warnings_df):
    all_errors_df = pd.concat([errors_df,warnings_df], ignore_index=True)
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
        all_errors_df.to_excel(writer, sheet_name='Validation Report', index=False)
def write_success_report(excel_path, warnings_df):
    success_row = [(0,"-","Data Uploded successfully","SUCCESS")]
    success_df = pd.DataFrame(success_row,columns=["row_number","column", "message","type"])
    succes_ans_warning = pd.concat([success_df,warnings_df],ignore_index=True)
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
        if not  succes_ans_warning.empty:
            succes_ans_warning.to_excel(writer, sheet_name='Validation Report',index=False)